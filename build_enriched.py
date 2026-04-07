import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load data
results_df = pd.read_csv('C:/Users/ethan/Downloads/international_results.csv', low_memory=False)
results_df['date'] = pd.to_datetime(results_df['date'])
teams_df = pd.read_excel('C:/Users/ethan/Downloads/fifa_wc2026_dataset.xlsx')

NAME_MAP = {'Curacaio': 'Cura\u00e7ao'}

# Build long-format results
def build_long(df):
    home = df.rename(columns={'home_team':'team','away_team':'opponent','home_score':'gf','away_score':'ga'}).copy()
    away = df.rename(columns={'away_team':'team','home_team':'opponent','away_score':'gf','home_score':'ga'}).copy()
    long = pd.concat([home, away], ignore_index=True)
    long['win']         = (long['gf'] > long['ga']).astype(int)
    long['draw']        = (long['gf'] == long['ga']).astype(int)
    long['clean_sheet'] = (long['ga'] == 0).astype(int)
    long['gd']          = long['gf'] - long['ga']
    return long

long = build_long(results_df)

# Recent form (Jan 2024 - May 2026, excl. WC 2026 matches)
form_df = long[
    (long['date'] >= '2024-01-01') &
    (long['date'] <= '2026-05-31') &
    (~((long['tournament'] == 'FIFA World Cup') & (long['date'] >= '2026-01-01')))
]

def form_stats(team):
    d = form_df[form_df['team'] == team]
    n = len(d)
    if n == 0:
        return {k: np.nan for k in ['form_games','form_win_rate','form_draw_rate',
                'form_goals_for_pg','form_goals_against_pg','form_clean_sheet_pct','form_gd_pg']}
    return {
        'form_games':            n,
        'form_win_rate':         round(d['win'].mean(), 3),
        'form_draw_rate':        round(d['draw'].mean(), 3),
        'form_goals_for_pg':     round(d['gf'].mean(), 2),
        'form_goals_against_pg': round(d['ga'].mean(), 2),
        'form_clean_sheet_pct':  round(d['clean_sheet'].mean(), 3),
        'form_gd_pg':            round(d['gd'].mean(), 2),
    }

# WC 2026 qualifying (2023 cycle)
qual_df = long[
    (long['tournament'] == 'FIFA World Cup qualification') &
    (long['date'] >= '2023-01-01') &
    (long['date'] <= '2026-03-31')
]

def qual_stats(team):
    d = qual_df[qual_df['team'] == team]
    n = len(d)
    if n == 0:
        return {k: np.nan for k in ['qual_games','qual_points_pg',
                'qual_goals_for_pg','qual_goals_against_pg','qual_clean_sheet_pct','qual_gd_pg']}
    pts = d['win'] * 3 + d['draw']
    return {
        'qual_games':            n,
        'qual_points_pg':        round(pts.mean(), 3),
        'qual_goals_for_pg':     round(d['gf'].mean(), 2),
        'qual_goals_against_pg': round(d['ga'].mean(), 2),
        'qual_clean_sheet_pct':  round(d['clean_sheet'].mean(), 3),
        'qual_gd_pg':            round(d['gd'].mean(), 2),
    }

# Last major tournament per confederation
CONF_TOURNEY = {
    'UEFA':     ('UEFA Euro',              '2024-06-01', '2024-07-31'),
    'CONMEBOL': ('Copa America',           '2024-06-01', '2024-07-31'),
    'CAF':      ('African Cup of Nations', '2023-12-01', '2024-02-29'),
    'AFC':      ('AFC Asian Cup',          '2024-01-01', '2024-02-29'),
    'CONCACAF': ('Gold Cup',               '2023-06-01', '2023-07-31'),
    'OFC':      ('Oceania Nations Cup',    '2024-01-01', '2024-12-31'),
}

def tourney_stats(team, confederation):
    empty = {'last_tourney': np.nan, 'last_tourney_games': np.nan,
             'last_tourney_gf': np.nan, 'last_tourney_ga': np.nan,
             'last_tourney_gd': np.nan, 'last_tourney_clean_sheets': np.nan}
    if confederation not in CONF_TOURNEY:
        return empty
    tourney, start, end = CONF_TOURNEY[confederation]
    if 'Copa' in tourney:
        mask = long['tournament'].str.contains('Copa', na=False)
    else:
        mask = (long['tournament'] == tourney)
    d = long[mask & (long['date'] >= start) & (long['date'] <= end) & (long['team'] == team)]
    n = len(d)
    if n == 0:
        return {**empty, 'last_tourney': tourney, 'last_tourney_games': 0}
    return {
        'last_tourney':              tourney,
        'last_tourney_games':        n,
        'last_tourney_gf':           int(d['gf'].sum()),
        'last_tourney_ga':           int(d['ga'].sum()),
        'last_tourney_gd':           int(d['gd'].sum()),
        'last_tourney_clean_sheets': int(d['clean_sheet'].sum()),
    }

# Build new columns for all 48 teams
records = []
for _, row in teams_df.iterrows():
    rname = NAME_MAP.get(row['team_name'], row['team_name'])
    r = {}
    r.update(form_stats(rname))
    r.update(qual_stats(rname))
    r.update(tourney_stats(rname, row['confederation']))
    records.append(r)

new_cols = pd.DataFrame(records)
enriched = pd.concat([teams_df, new_cols], axis=1)

# Save to new xlsx
out_path = 'C:/Users/ethan/Downloads/fifa_wc2026_enriched.xlsx'
enriched.to_excel(out_path, index=False, sheet_name='fifa_wc2026_dataset')

# Style: original headers in blue, new columns in light blue background
wb = load_workbook(out_path)
ws = wb.active

orig_cols = len(teams_df.columns)
new_col_fill     = PatternFill('solid', start_color='DDEEFF', end_color='DDEEFF')
orig_header_fill = PatternFill('solid', start_color='2E75B6', end_color='2E75B6')
new_header_fill  = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
white_font       = Font(bold=True, color='FFFFFF', name='Arial', size=10)
data_font        = Font(name='Arial', size=10)

for col_idx, cell in enumerate(ws[1], start=1):
    cell.fill = orig_header_fill if col_idx <= orig_cols else new_header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for row in ws.iter_rows(min_row=2):
    for col_idx, cell in enumerate(row, start=1):
        cell.font = data_font
        if col_idx > orig_cols:
            cell.fill = new_col_fill
        cell.alignment = Alignment(horizontal='center')

for col in ws.columns:
    max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 22)

ws.row_dimensions[1].height = 40
ws.freeze_panes = 'B2'
wb.save(out_path)

print('Saved:', out_path)
print('Final shape:', enriched.shape)
print('New columns added:', list(new_cols.columns))
print()
print('Sample - Brazil:')
print(enriched[enriched['team_name'] == 'Brazil'][list(new_cols.columns)].T.to_string())
print()
print('Sample - Spain:')
print(enriched[enriched['team_name'] == 'Spain'][list(new_cols.columns)].T.to_string())
